import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import statistics
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
    # 文件不存在，创建新的 Workbook
        workbook = Workbook()

    # 获取或创建一个工作表
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # 在指定列中插入数据
    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i+1)]
        cell.value = value

    # 保存文件
    workbook.save(excel_name)

def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, SF, SD, VIF, MS_SSIM

if __name__ == '__main__':
    with_mean = True
    # dataroot = 'E:/task/fusion/SKFusion/images/roadscene/'
    # results_root = 'E:/task/fusion/SKFusion/output_image_roadscene/'
    dataroot = 'E:\dataset/fusion/rgb/roadscene/'
    results_root = 'E:/task/fusion/SCFusion/output_image_roadscene/'
    #融合地址
    dataset = 'roadscene'
    ir_dir = os.path.join(dataroot, 'Inf')

    vi_dir = os.path.join(dataroot, 'Vis')
    # f_dir = os.path.join(results_root, dataset)
    f_dir = results_root
    save_dir = 'E:/task/fusion/SCFusion/output_image_roadscene'
    os.makedirs(save_dir, exist_ok=True)

    metric_save_name = os.path.join(save_dir, 'metric_{}.xlsx'.format(dataset))
    filelist = natsorted(os.listdir(ir_dir))

    Method_list = []
    for root, dirs, files in os.walk(save_dir):
        for dir_name in dirs:
            Method_list.append(dir_name)

    # Method_list = ['RFN-Nest_kd1', 'RFN-Nest_kd2', 'RFN-Nest_skfusion_5.0_0.5',
    #                'RFN-Nest_skfcafusion_6.0_3.0',
    #                'RFN-Nest_skfusion_6.0_3.0', 'RFN-Nest_skfusion_6.0_3.0_balance_loss']
    metric_list = ['', 'EN', 'MI', 'SF', 'SD', 'VIF', 'MS_SSIM']

    for i, Method in enumerate(Method_list):
        EN_list = []
        MI_list = []
        SF_list = []
        SD_list = []
        VIF_list = []
        MS_SSIM_list = []
        filename_list = ['']
        sub_f_dir = os.path.join(f_dir, Method)
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(sub_f_dir, item)
            EN, MI, SF, SD, VIF, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            SD_list.append(SD)
            VIF_list.append(VIF)
            MS_SSIM_list.append(MS_SSIM)
            filename_list.append(item)
            eval_bar.set_description("{} | {}".format(Method, item))
        if with_mean:
            # 添加均值
            EN_list.append(np.mean(EN_list))
            MI_list.append(np.mean(MI_list))
            SF_list.append(np.mean(SF_list))
            SD_list.append(np.mean(SD_list))
            VIF_list.append(np.mean(VIF_list))
            MS_SSIM_list.append(np.mean(MS_SSIM_list))
            filename_list.append('mean')

            ## 添加标准差
            EN_list.append(np.std(EN_list))
            MI_list.append(np.std(MI_list))
            SF_list.append(np.std(SF_list))
            SD_list.append(np.std(SD_list))
            VIF_list.append(np.std(VIF_list))
            MS_SSIM_list.append(np.std(MS_SSIM_list))
            filename_list.append('std')

        ## 保留三位小数
        EN_list = [round(x, 3) for x in EN_list]
        MI_list = [round(x, 3) for x in MI_list]
        SF_list = [round(x, 3) for x in SF_list]
        SD_list = [round(x, 3) for x in SD_list]
        VIF_list = [round(x, 3) for x in VIF_list]
        MS_SSIM_list = [round(x, 3) for x in MS_SSIM_list]

        EN_list.insert(0, '{}'.format(Method))
        MI_list.insert(0, '{}'.format(Method))
        SF_list.insert(0, '{}'.format(Method))
        SD_list.insert(0, '{}'.format(Method))
        VIF_list.insert(0, '{}'.format(Method))
        MS_SSIM_list.insert(0, '{}'.format(Method))

        mean_list = [EN_list[0], EN_list[-2], MI_list[-2], SF_list[-2],
                     SD_list[-2], VIF_list[-2],
                     MS_SSIM_list[-2]]
        MI_list_mean = [MI_list[0], MI_list[-2]]
        SF_list_mean = [SF_list[0], SF_list[-2]]
        SD_list_mean = [SD_list[0], SD_list[-2]]
        VIF_list_mean = [VIF_list[0], VIF_list[-2]]
        MS_SSIM_list_mean = [MS_SSIM_list[0], MS_SSIM_list[-2]]

        if i == 0:
            write_excel(metric_save_name, "total_metric", 0, metric_list)
            write_excel(metric_save_name, 'EN', 0, filename_list)
            write_excel(metric_save_name, "MI", 0, filename_list)
            write_excel(metric_save_name, "SF", 0, filename_list)
            write_excel(metric_save_name, "SD", 0, filename_list)
            write_excel(metric_save_name, "VIF", 0, filename_list)
            write_excel(metric_save_name, "MS_SSIM", 0, filename_list)

        write_excel(metric_save_name, 'total_metric', i + 1, mean_list)

        write_excel(metric_save_name, 'EN', i + 1, EN_list)
        write_excel(metric_save_name, 'MI', i + 1, MI_list)
        write_excel(metric_save_name, 'SF', i + 1, SF_list)
        write_excel(metric_save_name, 'SD', i + 1, SD_list)
        write_excel(metric_save_name, 'VIF', i + 1, VIF_list)
        write_excel(metric_save_name, 'MS_SSIM', i + 1, MS_SSIM_list)